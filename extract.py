from googleapiclient.discovery import build
from google.auth import default
from googleapiclient.http import MediaIoBaseDownload
import io
from openpyxl import load_workbook

def extract_data():
    creds, _ = default()
    service = build("drive", "v3", credentials=creds)

    print("\n✅ Autenticación correcta con la Service Account")

    SHARED_DRIVE_ID = "0APEWZBnOd0PhUk9PVA"
    PARENT_FOLDER_ID = "1nRvf2qydJ0G9pB8Yf9PCgfE_5ga9rEjq" 
    TARGET_FILENAME = "reporte_contable.xlsx"

    # ----------------------------
    # 1️⃣ Buscar el archivo dentro de la unidad compartida
    # ----------------------------
    print(f"\n🔍 Buscando '{TARGET_FILENAME}' en la unidad compartida...\n")

    results = service.files().list(
        q=(
            f"'{PARENT_FOLDER_ID}' in parents "
            f"and name='{TARGET_FILENAME}' "
            f"and mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' "
            f"and trashed = false"
        ),
        corpora="drive",
        driveId=SHARED_DRIVE_ID,
        includeItemsFromAllDrives=True,
        supportsAllDrives=True,
        fields="files(id, name, parents)"
    ).execute()

    files = results.get("files", [])

    if not files:
        print(f"⚠ No se encontró el archivo '{TARGET_FILENAME}'.")
        return

    file_id = files[0]["id"]
    print(f"✅ Archivo encontrado: {files[0]['name']} ({file_id})")

    # ----------------------------
    # 2️⃣ Descargar el archivo a memoria
    # ----------------------------
    print("\n⬇ Descargando archivo desde Google Drive...\n")

    request = service.files().get_media(fileId=file_id)
    fh = io.BytesIO()
    downloader = MediaIoBaseDownload(fh, request)

    done = False
    while not done:
        status, done = downloader.next_chunk()
        print(f"   Progreso: {int(status.progress() * 100)}%")

    fh.seek(0)

    # ----------------------------
    # 3️⃣ Leer el archivo con openpyxl
    # ----------------------------
    wb = load_workbook(fh, data_only=True)

    HOJAS_OBJETIVO = ["Alertas", "Cobradas dentro de los 60 días"]
    
    todas_las_filas = []

    # Recorremos las dos primeras hojas
    for nombre in HOJAS_OBJETIVO:
        if nombre not in wb.sheetnames:
            print(f"⚠ La hoja '{nombre}' no existe en el archivo")
            continue

        sheet = wb[nombre]
        print(f"\n📄 Procesando hoja: {sheet.title}")

        # Leer y pasar a minuscula encabezados
        headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        normalized = [str(h).strip().lower() if h is not None else "" for h in headers]

        if "a indyco" not in normalized:
            print("⚠ No se encontró la columna 'A indyco' en esta hoja.")
            continue

        indyco_col_idx = normalized.index("a indyco") + 1

        # Filtrar filas donde "A indyco" = "si"
        filter_rows = [
            row
            for row in sheet.iter_rows(min_row=2, values_only=True)
            if isinstance(row[indyco_col_idx - 1], str)
            and row[indyco_col_idx - 1].strip().lower() == "si"
        ]

        if not filter_rows:
            print("No se encontraron filas con el criterio 'si'")
        else:
            print(f"Se encontraron {len(filter_rows)} filas")

        todas_las_filas.extend(filter_rows)

    print(f"\n✅ Total de filas encontradas en ambas hojas: {len(todas_las_filas)}")
    return todas_las_filas