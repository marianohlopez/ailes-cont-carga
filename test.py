from googleapiclient.discovery import build
from google.auth import default
from googleapiclient.http import MediaIoBaseDownload
import io
from openpyxl import load_workbook

def main():
    creds, _ = default()
    service = build("drive", "v3", credentials=creds)

    print("\n✅ Autenticación correcta con la Service Account")

    DRIVE_ID = "0ALVObxz8zlW1Uk9PVA"  # ID de la unidad compartida
    TARGET_FILENAME = "reporte_contable.xlsx"

    # ----------------------------
    # 1️⃣ Buscar el archivo dentro de la unidad compartida
    # ----------------------------
    print(f"\n🔍 Buscando '{TARGET_FILENAME}' en la unidad compartida...\n")

    results = service.files().list(
        q=f"name='{TARGET_FILENAME}' and mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'",
        corpora="drive",
        driveId=DRIVE_ID,
        includeItemsFromAllDrives=True,
        supportsAllDrives=True,
        fields="files(id, name)"
    ).execute()

    files = results.get("files", [])

    if not files:
        print(f"⚠ No se encontró el archivo '{TARGET_FILENAME}'.")
        return

    file_id = files[0]["id"]
    print(f"✅ Archivo encontrado: {files[0]['name']} ({file_id})")

    # ----------------------------
    # 2️⃣ Descargar el archivo a memoria
    # ----------------------------
    print("\n⬇ Descargando archivo desde Google Drive...\n")

    request = service.files().get_media(fileId=file_id)
    fh = io.BytesIO()
    downloader = MediaIoBaseDownload(fh, request)

    done = False
    while not done:
        status, done = downloader.next_chunk()
        print(f"   Progreso: {int(status.progress() * 100)}%")

    fh.seek(0)

    # ----------------------------
    # 3️⃣ Leer el archivo con openpyxl
    # ----------------------------
    wb = load_workbook(fh, data_only=True)
    sheet = wb.active  # Toma la primera hoja

    # ----------------------------
    # 4️⃣ Buscar la columna "A indyco" y listar filas que digan SI
    # ----------------------------
    print("\n📊 Filtrando filas donde 'A indyco' = SI o si:\n")

    # Encontrar índice de la columna "A indyco"
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    normalized = [str(h).strip().lower() if h is not None else "" for h in headers]
    
    if "a indyco" not in normalized:
        print("⚠ No se encontró la columna 'A indyco' en la hoja.")
        return

    indyco_col_idx = normalized.index("a indyco") + 1

    # Iterar filas desde la segunda (los datos)
    found = False
    for row in sheet.iter_rows(min_row=2, values_only=True):
        indyco_value = row[indyco_col_idx - 1]
        if isinstance(indyco_value, str) and indyco_value.strip().lower() == "si":
            found = True
            print(row)

    if not found:
        print("⚠ No se encontraron filas con 'A indyco' = SI o si.")

if __name__ == "__main__":
    main()
